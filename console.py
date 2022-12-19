from openpyxl import Workbook, load_workbook
import pywhatkit

wb = load_workbook("phone.xlsx")
ws = wb.active
print(ws)
# saat ve dakikayı başladığınız zaman dilimine göre ayarlamalısınız.


for satir in range(1, ws.max_row + 1):
    hour = 15
    minute = 0
    print(str(ws.cell(satir, 1).value), end="")
    print()
    try:
        pywhatkit.sendwhatmsg(str(ws.cell(satir, 1).value), "Bu mesaj Pınar Mıhcıoğlu tarafından gönderilmiştir.",
                              hour + 1,
                              minute)
    except Exception as e:
        print("hata oluştu", e)
